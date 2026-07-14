import os
import uuid

from django.conf import settings
from django.http import FileResponse, Http404
from openpyxl import load_workbook
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView


def _upload_dir():
    path = settings.MEDIA_ROOT
    os.makedirs(path, exist_ok=True)
    return path


class UploadView(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded = request.FILES.get("file") or request.FILES.get("image")
        if not uploaded:
            return Response({"error": "No file provided"}, status=400)
        ext = os.path.splitext(uploaded.name)[1]
        filename = f"{uuid.uuid4().hex}{ext}"
        dest = os.path.join(_upload_dir(), filename)
        with open(dest, "wb+") as out:
            for chunk in uploaded.chunks():
                out.write(chunk)
        return Response({
            "filename": filename,
            "url": f"/uploads/{filename}",
            "fileInfo": {"filename": filename},
        })


class PreviewView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, filename):
        path = os.path.join(_upload_dir(), os.path.basename(filename))
        if not os.path.isfile(path):
            raise Http404("File not found")
        return FileResponse(open(path, "rb"))


class ReadExcelView(APIView):
    permission_classes = [AllowAny]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"error": "No file provided"}, status=400)
        wb = load_workbook(uploaded, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() for cell in row):
                rows.append([cell if cell is not None else "" for cell in row])
        return Response({"success": True, "data": rows})
